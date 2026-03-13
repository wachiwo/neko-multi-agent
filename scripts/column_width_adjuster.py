#!/usr/bin/env python3
"""
column_width_adjuster.py — Apply Excel 画面項目 column width specs to HTML data-table.

Reads display format/digits from Excel 画面項目 sheet, matches to HTML <th> headers,
and injects min-width/width CSS rules into the HTML <style> section.

Usage:
    python3 scripts/column_width_adjuster.py EXCEL_PATH HTML_PATH [--dry-run]
"""

import argparse
import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("ERROR: beautifulsoup4 is required. Install with: pip install beautifulsoup4", file=sys.stderr)
    sys.exit(1)

# ── Format-to-width mapping (from task spec / 001 reference) ──

FORMAT_WIDTH_MAP = [
    # (regex_pattern, css_property, px_value, description)
    (r'^#,###,###,##0$',         'min-width', 120, '10桁カンマ整数'),
    (r'^###,###,##0$',           'min-width', 100, '9桁カンマ整数'),
    (r'^###,###,##0\.0000$',     'min-width', 130, '小数4桁単価'),
    (r'^###,##0\.00$',           'min-width', 100, '小数2桁'),
    (r'^##0\.0+%$',              'width',      70, 'パーセント'),
    (r'^###0$',                  'width',      50, 'ランク/小整数'),
    (r'^##,##0$',                'min-width',  60, '5桁カウント'),
    (r'^yyyy/MM/dd$',            'min-width',  90, '日付'),
    (r'^yyyy/MM$',               'min-width',  70, '年月'),
]


def normalize_text(text):
    """NFKC normalize and strip whitespace for matching."""
    if text is None:
        return ''
    return unicodedata.normalize('NFKC', str(text)).strip()


def format_to_width(display_format, display_digits):
    """Determine CSS width from display format and digits.

    Returns: (css_property, px_value, description) or None
    """
    fmt = normalize_text(display_format)
    digits = display_digits

    # Try format pattern match first
    if fmt:
        for pattern, css_prop, px, desc in FORMAT_WIDTH_MAP:
            if re.match(pattern, fmt):
                return (css_prop, px, desc)

    # Fallback: text/code types based on digits
    if digits is not None:
        try:
            d = int(digits)
        except (ValueError, TypeError):
            return None

        if d <= 0:
            return None

        # Short codes (<=6 digits): fixed width
        if d <= 6 and not fmt:
            return ('width', max(40, d * 10), f'{d}桁コード')

        # Text fields: 8px per character, cap at 200px
        if not fmt or fmt == '':
            px = max(60, min(200, d * 8))
            return ('min-width', px, f'{d}桁テキスト')

    return None


def read_excel_columns(excel_path):
    """Read grid column definitions from Excel 画面項目 sheet.

    Returns: list of dicts with keys:
        name, digits, format, align, section
    """
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)

    # Find 画面項目 sheet
    target_sheet = None
    for name in wb.sheetnames:
        normalized = normalize_text(name)
        if normalized == '画面項目':
            target_sheet = name
            break
    if target_sheet is None:
        # Try partial match
        for name in wb.sheetnames:
            if '画面項目' in normalize_text(name) and '✕' not in name:
                target_sheet = name
                break

    if target_sheet is None:
        wb.close()
        return []

    ws = wb[target_sheet]

    # Find header row (contains 'n°' or '表示項目')
    header_row = None
    col_map = {}
    for row_idx in range(1, 15):
        for col_idx in range(1, 45):
            val = normalize_text(ws.cell(row=row_idx, column=col_idx).value)
            if val in ('n°', 'No', 'no', 'NO', 'n'):
                col_map['no'] = col_idx
            elif val == '表示項目':
                col_map['name'] = col_idx
            elif val == 'コントロール':
                col_map['control'] = col_idx
            elif val == 'I/O':
                col_map['io'] = col_idx
            elif val == '表示桁数':
                col_map['digits'] = col_idx
            elif val == '表示位置':
                col_map['align'] = col_idx
            elif val == '表示形式':
                col_map['format'] = col_idx
        if 'name' in col_map and 'control' in col_map:
            header_row = row_idx
            break

    if header_row is None:
        wb.close()
        return []

    # Read data rows
    columns = []
    current_section = None
    max_row = ws.max_row or 200

    for row_idx in range(header_row + 1, min(max_row + 1, 500)):
        name_val = normalize_text(ws.cell(row=row_idx, column=col_map.get('name', 4)).value)
        ctrl_val = normalize_text(ws.cell(row=row_idx, column=col_map.get('control', 12)).value)
        no_val = ws.cell(row=row_idx, column=col_map.get('no', 2)).value

        if not name_val:
            continue

        # Section header: has name but no n° number
        if no_val is None and not ctrl_val:
            current_section = name_val
            continue

        # Only grid items are table columns
        # Convention 1: ctrl_val contains 'グリッド' directly
        # Convention 2: item is within a section whose name contains 'グリッド'
        is_grid_ctrl = 'グリッド' in ctrl_val
        is_in_grid_section = current_section is not None and 'グリッド' in current_section
        if not is_grid_ctrl and not is_in_grid_section:
            continue

        digits_val = ws.cell(row=row_idx, column=col_map.get('digits', 20)).value
        fmt_val = ws.cell(row=row_idx, column=col_map.get('format', 35)).value
        align_val = ws.cell(row=row_idx, column=col_map.get('align', 28)).value

        columns.append({
            'name': name_val,
            'digits': digits_val,
            'format': normalize_text(fmt_val) if fmt_val else None,
            'align': normalize_text(align_val) if align_val else None,
            'section': current_section,
        })

    wb.close()
    return columns


def extract_th_with_positions(html_content):
    """Extract <th> texts from data-tables with correct visual column positions.
    Handles multi-row headers with rowspan/colspan.

    Returns: list of lists of (th_text, visual_col_1based) per table.
    """
    soup = BeautifulSoup(html_content, 'html.parser')
    tables = soup.find_all('table', class_=re.compile(r'data-table'))
    result = []

    for table in tables:
        thead = table.find('thead')
        if not thead:
            continue
        header_rows = thead.find_all('tr')
        if not header_rows:
            continue
        num_rows = len(header_rows)

        if num_rows == 1:
            ths = header_rows[0].find_all('th')
            result.append([(normalize_text(th.get_text()), i + 1) for i, th in enumerate(ths)])
            continue

        # Multi-row header: build occupancy grid for rowspan/colspan
        occupied = set()
        th_positions = {}  # (row_idx, th_dom_idx) -> visual_col

        for row_idx, tr in enumerate(header_rows):
            col = 1
            for th_idx, th in enumerate(tr.find_all('th')):
                while (row_idx, col) in occupied:
                    col += 1
                rowspan = int(th.get('rowspan', 1))
                colspan = int(th.get('colspan', 1))
                th_positions[(row_idx, th_idx)] = col
                for r in range(rowspan):
                    for c in range(colspan):
                        occupied.add((row_idx + r, col + c))
                col += colspan

        last_row = num_rows - 1
        th_data = []

        # Include upper-row ths with rowspan spanning to bottom (single-column headers)
        for row_idx in range(last_row):
            for th_idx, th in enumerate(header_rows[row_idx].find_all('th')):
                rowspan = int(th.get('rowspan', 1))
                colspan = int(th.get('colspan', 1))
                if row_idx + rowspan > last_row and colspan == 1:
                    visual_col = th_positions[(row_idx, th_idx)]
                    th_data.append((normalize_text(th.get_text()), visual_col))

        # Bottom row ths
        for th_idx, th in enumerate(header_rows[last_row].find_all('th')):
            visual_col = th_positions.get((last_row, th_idx), th_idx + 1)
            th_data.append((normalize_text(th.get_text()), visual_col))

        th_data.sort(key=lambda x: x[1])
        result.append(th_data)

    return result


def match_columns_to_headers(excel_cols, th_data_list):
    """Match Excel column definitions to HTML th headers.

    th_data_list: list of lists of (th_text, visual_col) per table.
    Returns: list of match dicts (deduplicated across tables).
    """
    matches = []
    seen_rules = set()

    for th_data in th_data_list:
        used_excel = set()

        for th_text, visual_col in th_data:
            if not th_text:
                continue

            best_match = None
            for i, ecol in enumerate(excel_cols):
                if i in used_excel:
                    continue
                if normalize_text(ecol['name']) == th_text:
                    best_match = i
                    break

            if best_match is None:
                for i, ecol in enumerate(excel_cols):
                    if i in used_excel:
                        continue
                    ecol_name = normalize_text(ecol['name'])
                    if ecol_name and (ecol_name in th_text or th_text in ecol_name):
                        best_match = i
                        break

            if best_match is not None:
                ecol = excel_cols[best_match]
                width_info = format_to_width(ecol['format'], ecol['digits'])
                if width_info:
                    css_prop, px, desc = width_info
                    rule_key = (visual_col, css_prop, px)
                    if rule_key not in seen_rules:
                        seen_rules.add(rule_key)
                        matches.append({
                            'th_index': visual_col,
                            'name': ecol['name'],
                            'css_prop': css_prop,
                            'px': px,
                            'desc': desc,
                            'format': ecol['format'],
                            'digits': ecol['digits'],
                            'align': ecol['align'],
                        })
                used_excel.add(best_match)

    return matches


def generate_css_block(matches):
    """Generate CSS rules for matched columns.

    Returns: CSS string to inject.
    """
    if not matches:
        return ''

    lines = []
    lines.append('')
    lines.append('/* Excel画面項目準拠の列幅 */')

    for m in matches:
        idx = m['th_index']
        name = m['name']
        fmt = m['format'] or ''
        desc = m['desc']
        css_prop = m['css_prop']
        px = m['px']

        comment = f'/* {name}: {desc}'
        if fmt:
            comment += f' ({fmt})'
        comment += ' */'

        lines.append(comment)
        lines.append(f'.data-table th:nth-child({idx}),')
        lines.append(f'.data-table td:nth-child({idx}) {{')
        lines.append(f'    {css_prop}: {px}px;')

        # Add text-align based on Excel 表示位置
        align = m.get('align')
        if align:
            if '右' in align:
                lines.append('    text-align: right;')
            elif '左' in align:
                lines.append('    text-align: left;')
            elif '中央' in align or '中' in align:
                lines.append('    text-align: center;')

        lines.append('}')

    return '\n'.join(lines)


def generate_common_css():
    """Generate common CSS rules for data-table (nowrap + padding)."""
    return """
/* 共通: white-space:nowrap + padding */
.data-table th,
.data-table td {
    white-space: nowrap;
    padding: 8px 10px;
}"""


def inject_css_into_html(html_content, css_block, common_css):
    """Inject column width CSS into HTML <style> section.

    Inserts before </style>. If existing 'Excel画面項目準拠の列幅' block exists, replaces it.
    """
    # Check if already has column width section — replace it
    pattern_existing = r'/\* Excel画面項目準拠の列幅 \*/.*?(?=\n/\*[^E]|\n\s*</style>|\n\s*\n/\*|\Z)'
    if re.search(r'/\* Excel画面項目準拠の列幅 \*/', html_content):
        # Remove old block: from the comment to the next non-column-width CSS block
        html_content = re.sub(
            r'\n/\* Excel画面項目準拠の列幅 \*/\n(?:.*?\n)*?(?=\n/\*(?! Excel| 共通:)|</style>)',
            '\n',
            html_content,
            flags=re.DOTALL,
        )
        # Also remove old common CSS if present
        html_content = re.sub(
            r'\n/\* 共通: white-space:nowrap \+ padding \*/\n(?:.*?\n)*?(?=\n/\*|</style>)',
            '\n',
            html_content,
            flags=re.DOTALL,
        )

    # Check if common rules already exist (white-space: nowrap on .data-table th/td)
    has_nowrap_th = bool(re.search(
        r'\.data-table\s+th[^{]*\{[^}]*white-space\s*:\s*nowrap',
        html_content, re.DOTALL
    ))
    has_nowrap_td = bool(re.search(
        r'\.data-table\s+td[^{]*\{[^}]*white-space\s*:\s*nowrap',
        html_content, re.DOTALL
    ))

    inject = ''
    if not (has_nowrap_th and has_nowrap_td):
        inject += common_css
    inject += css_block + '\n'

    # Find </style> and inject before it
    style_end = html_content.rfind('</style>')
    if style_end == -1:
        print("WARNING: No </style> found in HTML", file=sys.stderr)
        return html_content

    return html_content[:style_end] + inject + '\n' + html_content[style_end:]


def main():
    parser = argparse.ArgumentParser(description='Apply Excel column width specs to HTML data-table')
    parser.add_argument('excel_path', help='Path to Excel design spec file (.xlsx)')
    parser.add_argument('html_path', help='Path to HTML prototype file')
    parser.add_argument('--dry-run', action='store_true', help='Show CSS without modifying HTML')
    args = parser.parse_args()

    excel_path = Path(args.excel_path)
    html_path = Path(args.html_path)

    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)
    if not html_path.exists():
        print(f"ERROR: HTML file not found: {html_path}", file=sys.stderr)
        sys.exit(1)

    # Step 1: Read Excel columns
    print(f"Reading Excel: {excel_path.name}")
    excel_cols = read_excel_columns(excel_path)
    if not excel_cols:
        print("WARNING: No grid columns found in 画面項目 sheet", file=sys.stderr)
        sys.exit(0)
    print(f"  Found {len(excel_cols)} grid columns:")
    for c in excel_cols:
        print(f"    {c['name']} | digits={c['digits']} | format={c['format']} | align={c['align']}")

    # Step 2: Read HTML headers
    print(f"\nReading HTML: {html_path.name}")
    html_content = html_path.read_text(encoding='utf-8')
    th_data_list = extract_th_with_positions(html_content)
    if not th_data_list:
        print("WARNING: No data-table <thead> found in HTML", file=sys.stderr)
        sys.exit(0)
    for i, th_data in enumerate(th_data_list):
        print(f"  Table {i+1} headers: {[(t, f'col{c}') for t, c in th_data]}")

    # Step 3: Match columns to headers
    matches = match_columns_to_headers(excel_cols, th_data_list)
    print(f"\nMatched {len(matches)} columns:")
    for m in matches:
        print(f"  th:{m['th_index']} '{m['name']}' → {m['css_prop']}: {m['px']}px ({m['desc']})")

    unmatched_excel = []
    matched_names = {m['name'] for m in matches}
    for c in excel_cols:
        if c['name'] not in matched_names:
            unmatched_excel.append(c['name'])
    if unmatched_excel:
        print(f"\n  Unmatched Excel columns: {unmatched_excel}")

    # Step 4: Generate CSS
    css_block = generate_css_block(matches)
    common_css = generate_common_css()

    if args.dry_run:
        print("\n=== DRY-RUN: CSS to inject ===")
        print(common_css)
        print(css_block)
        print("=== END DRY-RUN ===")
        return

    # Step 5: Inject CSS into HTML
    new_html = inject_css_into_html(html_content, css_block, common_css)

    # Step 6: Write back
    html_path.write_text(new_html, encoding='utf-8')
    print(f"\nCSS injected into: {html_path}")
    print("Done.")


if __name__ == '__main__':
    main()
