#!/usr/bin/env python3
"""
Excel→HTML Field Width Adjuster
Based on neko-excel-html-field-width-adjuster skill.
Reads Excel 画面項目 sheet, extracts field specs, matches to HTML inputs, applies width+maxlength.
"""

import re
import sys
import unicodedata
import openpyxl

# === CONFIG ===
WIDTH_RULES = {
    (1, 3): 80,
    (4, 6): 120,
    (7, 10): 180,
    (11, 20): 250,
    (21, 50): 400,
}
DATE_WIDTH = 160

def normalize(text):
    if text is None:
        return ""
    return unicodedata.normalize('NFKC', str(text)).strip()

def calc_width(digits):
    for (lo, hi), px in WIDTH_RULES.items():
        if lo <= digits <= hi:
            return px
    return None

def extract_fields(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheets = [s for s in wb.sheetnames if '画面項目' in s and '✕' not in s]
    fields = []

    for sheet_name in sheets:
        ws = wb[sheet_name]
        header_row = 5
        col_map = {'name': 4, 'control': 12, 'io': 18, 'display_digits': 20, 'input_digits': 24}

        for r in range(1, 10):
            for c in range(1, 45):
                val = normalize(ws.cell(row=r, column=c).value)
                if val == '表示項目':
                    col_map['name'] = c
                    header_row = r
                elif val == 'コントロール':
                    col_map['control'] = c
                elif val == 'I/O':
                    col_map['io'] = c
                elif val == '表示桁数':
                    col_map['display_digits'] = c
                elif val == '入力桁数':
                    col_map['input_digits'] = c

        max_row = ws.max_row or 200
        for row_idx in range(header_row + 1, min(max_row + 1, 1100)):
            name = normalize(ws.cell(row=row_idx, column=col_map['name']).value)
            control = normalize(ws.cell(row=row_idx, column=col_map['control']).value)
            io = normalize(ws.cell(row=row_idx, column=col_map['io']).value)
            display_d = ws.cell(row=row_idx, column=col_map['display_digits']).value
            input_d = ws.cell(row=row_idx, column=col_map['input_digits']).value

            if not name:
                continue
            if 'テキストボックス' not in control and '日付ピッカー' not in control:
                continue
            if io == 'O':
                continue

            is_date = '日付' in control
            digits = input_d or display_d
            if digits is None and not is_date:
                continue

            fields.append({
                'name': name,
                'digits': int(digits) if digits else 10,
                'control': control,
                'is_date': is_date,
            })

    wb.close()
    return fields

def strip_html_tags(text):
    text = re.sub(r'<span[^>]*class="required"[^>]*>\*?</span>', '', text)
    text = re.sub(r'<[^>]+>', '', text)
    return normalize(text)

def find_label_positions(html):
    labels = []
    for m in re.finditer(r'<label[^>]*>(.*?)</label>', html, re.DOTALL):
        text = strip_html_tags(m.group(1))
        labels.append({'text': text, 'start': m.start(), 'end': m.end()})
    return labels

def find_inputs_in_range(html, start, end):
    segment = html[start:end]
    inputs = []
    for m in re.finditer(r'<input\b[^>]*>', segment):
        abs_start = start + m.start()
        abs_end = start + m.end()
        tag = m.group(0)
        type_match = re.search(r'type="([^"]*)"', tag)
        input_type = type_match.group(1) if type_match else 'text'
        inputs.append({
            'tag': tag, 'start': abs_start, 'end': abs_end,
            'type': input_type,
        })
    return inputs

def is_separator(text):
    return text in ('～', '~', '以上', '以下', '−', '-', '', '〜')

def match_field_to_label(field, labels):
    name = field['name']

    from_to = None
    ft_match = re.search(r'[（(](FROM|TO|from|to|From|To)[）)]', name)
    if ft_match:
        from_to = ft_match.group(1).upper()
        name = re.sub(r'[（(](FROM|TO|from|to|From|To)[）)]', '', name).strip()

    if '.' in name:
        name = name.split('.')[0]

    for label in labels:
        lt = label['text']
        if is_separator(lt):
            continue
        if name == lt:
            return label, from_to

    for label in labels:
        lt = label['text']
        if is_separator(lt):
            continue
        if len(name) >= 2 and len(lt) >= 2:
            if name.startswith(lt) or lt.startswith(name):
                return label, from_to

    return None, from_to

def apply_width_to_tag(tag, width_px):
    width_re = re.compile(r'(?<![a-z-])width\s*:\s*[^;]+;?\s*')

    if 'style="' in tag:
        style_match = re.search(r'style="([^"]*)"', tag)
        if style_match:
            old_style = style_match.group(1)
            new_style = width_re.sub('', old_style).strip().rstrip(';')
            if new_style:
                new_style = f"width: {width_px}px; {new_style}"
            else:
                new_style = f"width: {width_px}px"
            return tag.replace(f'style="{old_style}"', f'style="{new_style}"')
    else:
        return tag.replace('<input ', f'<input style="width: {width_px}px" ', 1)
    return tag

def _insert_attr(tag, attr_str):
    """Insert attribute before closing /> or >. Handles self-closing tags."""
    if re.search(r'/\s*>\s*$', tag):
        return re.sub(r'\s*/\s*>\s*$', f' {attr_str} />', tag)
    return re.sub(r'>\s*$', f' {attr_str}>', tag)

def apply_maxlength_to_tag(tag, digits, input_type):
    if input_type in ('date', 'month'):
        return tag

    if input_type == 'number':
        max_val = 10**digits - 1
        if re.search(r'\bmax="', tag):
            tag = re.sub(r'\bmax="[^"]*"', f'max="{max_val}"', tag)
        else:
            tag = _insert_attr(tag, f'max="{max_val}"')

    if 'maxlength=' in tag:
        tag = re.sub(r'maxlength="[^"]*"', f'maxlength="{digits}"', tag)
    else:
        tag = _insert_attr(tag, f'maxlength="{digits}"')

    return tag

def process_file(excel_path, html_path):
    print(f"Reading Excel: {excel_path.split('/')[-1]}")
    fields = extract_fields(excel_path)
    print(f"  Extracted {len(fields)} input fields from Excel")

    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()
    original_html = html

    labels = find_label_positions(html)
    modifications = []
    matched_count = 0
    unmatched = []

    for field in fields:
        label, from_to = match_field_to_label(field, labels)
        if label is None:
            unmatched.append(field['name'])
            continue

        div_end = html.find('</div>', label['end'])
        if div_end == -1:
            div_end = len(html)
        inputs = find_inputs_in_range(html, label['end'], div_end)

        if not inputs:
            unmatched.append(f"{field['name']}(no-input)")
            continue

        inputs = [i for i in inputs if 'type="hidden"' not in i['tag']]

        if from_to == 'FROM' and len(inputs) >= 1:
            target_inputs = [inputs[0]]
        elif from_to == 'TO' and len(inputs) >= 2:
            target_inputs = [inputs[1]]
        elif from_to == 'TO' and len(inputs) == 1:
            target_inputs = [inputs[0]]
        else:
            target_inputs = inputs

        for inp in target_inputs:
            is_date = inp['type'] in ('date', 'month') or field.get('is_date')
            width = DATE_WIDTH if is_date else calc_width(field['digits'])
            new_tag = inp['tag']

            if is_date:
                if width is not None:
                    new_tag = apply_width_to_tag(new_tag, width)
            else:
                if width is not None:
                    new_tag = apply_width_to_tag(new_tag, width)
                new_tag = apply_maxlength_to_tag(new_tag, field['digits'], inp['type'])

            if new_tag != inp['tag']:
                modifications.append((inp['start'], inp['end'], inp['tag'], new_tag))
                matched_count += 1

    # Apply back-to-front
    modifications.sort(key=lambda m: m[0], reverse=True)
    for start, end, old_tag, new_tag in modifications:
        html = html[:start] + new_tag + html[end:]

    # Corruption check
    corruption = re.findall(r'>[^<]{0,5}(?:type=|value=|maxlength=|style=|placeholder=)', html)
    corruption_count = len(corruption) if corruption else 0
    if corruption:
        print(f"  WARNING: Corruption detected: {corruption[:3]}")

    # JS unchanged
    orig_scripts = len(re.findall(r'<script\b', original_html, re.IGNORECASE))
    new_scripts = len(re.findall(r'<script\b', html, re.IGNORECASE))
    if orig_scripts != new_scripts:
        print(f"  WARNING: Script block count changed! {orig_scripts} → {new_scripts}")

    # Date+maxlength check
    date_ml = re.findall(r'<input[^>]*type="(?:date|month)"[^>]*maxlength', html)
    date_ml2 = re.findall(r'<input[^>]*maxlength[^>]*type="(?:date|month)"', html)
    if date_ml or date_ml2:
        print(f"  WARNING: maxlength on date/month: {len(date_ml) + len(date_ml2)}")

    # Nonepx check
    nonepx = html.count('Nonepx')
    if nonepx:
        print(f"  WARNING: Nonepx artifacts: {nonepx}")

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"  Result: {matched_count} matched, {len(unmatched)} unmatched, {corruption_count} corruption")
    if unmatched:
        shown = unmatched[:15]
        print(f"  Unmatched: {', '.join(shown)}")
        if len(unmatched) > 15:
            print(f"    ... +{len(unmatched) - 15} more")

    return matched_count, len(unmatched), corruption_count

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python3 field_width_adjuster.py <excel_path> <html_path>")
        sys.exit(1)
    m, u, c = process_file(sys.argv[1], sys.argv[2])
    sys.exit(1 if c > 0 else 0)
